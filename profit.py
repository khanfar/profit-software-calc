import csv
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment

def load_employee_data(file_path):
    employee_data = {}
    with open(file_path, encoding='utf-8') as file:
        for line in file:
            name, salary, date = line.strip().split(',')
            date = datetime.strptime(date, '%Y-%m-%d')
            salary = float(salary)
            
            if name not in employee_data:
                employee_data[name] = []
                
            employee_data[name].append((date, salary))
            
    return employee_data

def load_your_data(file_path):
    entries = []
    with open(file_path, encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            if len(row) > 0:
                entries.append(row)
    return entries

def calculate_monthly_income(entries, month, year):
    income_data = {}
    for entry in entries:
        try:
            vehicle_plate, vehicle_type, entry_date, company, driver, phone, mechanic, note1, note2, final_report = entry
            entry_date = datetime.strptime(entry_date, '%d.%m.%Y')

            if entry_date.month == month and entry_date.year == year:
                mechanic = mechanic.strip()
                if 'شيكل' in final_report:
                    amount_str = final_report.split('شيكل')[0].strip().split()[-1]
                    if 'شغل' in final_report:
                        amount_str = final_report.split('شغل')[1].split('شيكل')[0].strip()
                    amount = int(amount_str)
                    if mechanic:
                        if mechanic not in income_data:
                            income_data[mechanic] = 0
                        income_data[mechanic] += amount
        except Exception as e:
            continue
    return income_data

def calculate_range_income(entries, start_date, end_date):
    income_data = {}
    for entry in entries:
        try:
            vehicle_plate, vehicle_type, entry_date, company, driver, phone, mechanic, note1, note2, final_report = entry
            entry_date = datetime.strptime(entry_date, '%d.%m.%Y')

            if start_date <= entry_date <= end_date:
                mechanic = mechanic.strip()
                if 'شيكل' in final_report:
                    amount_str = final_report.split('شيكل')[0].strip().split()[-1]
                    if 'شغل' in final_report:
                        amount_str = final_report.split('شغل')[1].split('شيكل')[0].strip()
                    amount = int(amount_str)
                    if mechanic:
                        if mechanic not in income_data:
                            income_data[mechanic] = 0
                        income_data[mechanic] += amount
        except Exception as e:
            continue
    return income_data

def generate_report(employee_data, income_data, month, year, output_file_txt, output_file_xlsx):
    with open(output_file_txt, 'w', encoding='utf-8') as file:
        # Write month and year at the top
        file.write(f"Report for {year}-{month:02d}\n")
        file.write("="*40 + "\n")
        
        for employee, salary_entries in employee_data.items():
            total_salary = sum(salary for date, salary in salary_entries if date.month == month and date.year == year)
            total_income = income_data.get(employee, 0)
            profit = total_income - total_salary
            file.write(f"{employee}, Total Salary: {total_salary}, Total Income: {total_income}, Profit: {profit}\n")

    # Excel report generation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {year}-{month:02d}"
    
    # Write report title
    ws['A1'] = f"Report for {year}-{month:02d}"
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Writing headers in Arabic
    ws.append(["موظف", "الراتب الإجمالي", "الدخل الإجمالي", "الأرباح"])
    
    # Writing data
    for employee, salary_entries in employee_data.items():
        total_salary = sum(salary for date, salary in salary_entries if date.month == month and date.year == year)
        total_income = income_data.get(employee, 0)
        profit = total_income - total_salary
        ws.append([employee, total_salary, total_income, profit])
    
    wb.save(output_file_xlsx)

def generate_report_range(employee_data, income_data, start_date, end_date, output_file_txt, output_file_xlsx):
    with open(output_file_txt, 'w', encoding='utf-8') as file:
        # Write date range at the top
        file.write(f"Report for {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}\n")
        file.write("="*40 + "\n")
        
        for employee, salary_entries in employee_data.items():
            total_salary = sum(salary for date, salary in salary_entries if start_date <= date <= end_date)
            total_income = income_data.get(employee, 0)
            profit = total_income - total_salary
            file.write(f"{employee}, Total Salary: {total_salary}, Total Income: {total_income}, Profit: {profit}\n")

    # Excel report generation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    
    # Write report title
    ws['A1'] = f"Report for {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Writing headers in Arabic
    ws.append(["موظف", "الراتب الإجمالي", "الدخل الإجمالي", "الأرباح"])
    
    # Writing data
    for employee, salary_entries in employee_data.items():
        total_salary = sum(salary for date, salary in salary_entries if start_date <= date <= end_date)
        total_income = income_data.get(employee, 0)
        profit = total_income - total_salary
        ws.append([employee, total_salary, total_income, profit])
    
    wb.save(output_file_xlsx)

def main():
    employee_file = 'employee_data.txt'
    your_data_file = 'your_data.csv'
    output_dir = 'output'
    
    # Ensure the output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    choice = input("Do you want to specify a single month (1) or a range of dates (2)? Enter 1 or 2: ")

    if choice == '1':
        month = int(input("Enter the month (1-12): "))
        year = int(input("Enter the year (e.g., 2024): "))
        output_file_txt = os.path.join(output_dir, f'report_{year}_{month:02d}.txt')
        output_file_xlsx = os.path.join(output_dir, f'report_{year}_{month:02d}.xlsx')
        
        employee_data = load_employee_data(employee_file)
        your_data_entries = load_your_data(your_data_file)
        income_data = calculate_monthly_income(your_data_entries, month, year)
        generate_report(employee_data, income_data, month, year, output_file_txt, output_file_xlsx)
    
    elif choice == '2':
        start_date_str = input("Enter the start date (YYYY-MM-DD): ")
        end_date_str = input("Enter the end date (YYYY-MM-DD): ")
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        output_file_txt = os.path.join(output_dir, f'report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.txt')
        output_file_xlsx = os.path.join(output_dir, f'report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx')
        
        employee_data = load_employee_data(employee_file)
        your_data_entries = load_your_data(your_data_file)
        income_data = calculate_range_income(your_data_entries, start_date, end_date)
        generate_report_range(employee_data, income_data, start_date, end_date, output_file_txt, output_file_xlsx)
    
    else:
        print("Invalid choice. Please run the program again and enter either 1 or 2.")

if __name__ == "__main__":
    main()
